import os
from openpyxl import load_workbook
from PIL import Image
from io import BytesIO
from tkinter import Tk, filedialog

def extract_floating_images(excel_path, output_folder):
    wb = load_workbook(excel_path)
    ws = wb.active

    os.makedirs(output_folder, exist_ok=True)

    image_count = 0
    for image in ws._images:
        try:
            anchor = image.anchor._from
        except AttributeError:
            print("❌ Couldn't find anchor position for an image.")
            continue

        row_index = anchor.row + 1
        col_index = anchor.col + 1

        if col_index != 3:
            continue  # Only process images in Column C

        image_name_cell = ws.cell(row=row_index, column=1)
        image_name = image_name_cell.value

        if not image_name:
            print(f"⚠️  Row {row_index}: No image name found in column A.")
            continue

        try:
            img_data = image._data()
            img = Image.open(BytesIO(img_data))
            output_path = os.path.join(output_folder, image_name)
            img.save(output_path)
            print(f"✅ Saved: {output_path}")
            image_count += 1
        except Exception as e:
            print(f"❌ Failed to save image at row {row_index}: {e}")

    if image_count == 0:
        print("⚠️  No matching images were found or saved.")
    else:
        print(f"\n🎉 Done! {image_count} image(s) saved to: {output_folder}")

def main():
    Tk().withdraw()
    excel_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel files", "*.xlsx")]
    )

    if not excel_path:
        print("❌ No file selected.")
        return

    # ✅ Set fixed output folder path
    output_folder = r"C:\Users\gudav\Downloads\Test"
    extract_floating_images(excel_path, output_folder)

if __name__ == "__main__":
    main()
