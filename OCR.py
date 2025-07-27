import os
import zipfile
import shutil
from flask import Flask, render_template,send_file, request
import pandas as pd
import easyocr
from openpyxl import load_workbook
from io import BytesIO
from PIL import Image
import xlsxwriter

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'static/images'

# Absolute temp directories for inspection
TEMP_DIR_1 = os.path.join('static', 'temp', 'temp1')
TEMP_DIR_2 = os.path.join('static', 'temp', 'temp2')

os.makedirs(TEMP_DIR_1, exist_ok=True)
os.makedirs(TEMP_DIR_2, exist_ok=True)

def extract_zip(zip_file, temp_dir):
    if os.path.exists(temp_dir):
        shutil.rmtree(temp_dir)
    os.makedirs(temp_dir, exist_ok=True)
    with zipfile.ZipFile(zip_file, 'r') as zip_ref:
        zip_ref.extractall(temp_dir)
    return temp_dir

def get_excel_files(folder):
    excel_files = []
    for root, _, files in os.walk(folder):
        for f in files:
            if f.lower().endswith(('.xlsx', '.xls')):
                excel_files.append(os.path.join(root, f))
    return excel_files

def read_image_names_from_excel(excel_path):
    df = pd.read_excel(excel_path, header=None)
    df = df.iloc[8:]  # Skip first 8 rows
    df = df.reset_index(drop=True)
    image_names = []
    for _, row in df.iterrows():
        if pd.notna(row[0]):  # Read from Column A
            name = str(row[0]).strip().replace("#", "")
            if name.endswith(".png"):
                image_names.append(name)
    return image_names

def extract_images_from_excel(excel_file, output_folder):
    wb = load_workbook(excel_file)
    ws = wb.active
    image_map = {}

    for img in ws._images:
        anchor = getattr(img.anchor, '_from', getattr(img.anchor, 'from_', None))
        if anchor:
            row = anchor.row + 1
            image_map[row] = img

    extracted = []
    for row in ws.iter_rows(min_row=9):
        img_name = str(row[0].value).strip().replace("#", "").replace("\t", "").replace(" ", "_") if row[0].value else None
        row_idx = row[0].row

        if img_name and row_idx in image_map and img_name.endswith(".png"):
            img = image_map[row_idx]
            img_bytes = img._data() if callable(img._data) else img._data
            if isinstance(img_bytes, BytesIO):
                image = Image.open(img_bytes)
            else:
                image = Image.open(BytesIO(img_bytes))

            save_path = os.path.join(output_folder, img_name)
            os.makedirs(os.path.dirname(save_path), exist_ok=True)
            image.save(save_path)
            extracted.append((img_name, save_path))

    return extracted

def perform_ocr(image_path, language_code):
    try:
        reader = easyocr.Reader([language_code], gpu=False)
        result = reader.readtext(image_path, detail=0)
        return ' '.join(result)
    except Exception as e:
        return f"OCR Failed: {str(e)}"

def find_file_by_name(root_dir, filename):
    for root, _, files in os.walk(root_dir):
        for file in files:
            if file.lower() == filename.lower():
                return os.path.join(root, file)
    return None

def load_languages():
    lang_path = os.path.join(app.root_path, 'static', 'css', 'language.txt')
    languages = []
    with open(lang_path, 'r', encoding='utf-8') as f:
        for line in f:
            if line.strip():
                parts = line.strip().split('\t')
                if len(parts) == 2:
                    code, name = parts
                    languages.append({'code': code.strip(), 'name': name.strip()})
    return languages


@app.route('/')
def home():
    languages = load_languages()  # Load from language.txt
    return render_template("home.html", languages=languages)

@app.route('/process', methods=['POST'])
def process():
    zip1 = request.files['zip_file_1']
    zip2 = request.files['zip_file_2']
    language_code = request.form['language']

    temp_dir1 = extract_zip(zip1, TEMP_DIR_1)
    temp_dir2 = extract_zip(zip2, TEMP_DIR_2)

    excels1 = get_excel_files(temp_dir1)
    excels2 = get_excel_files(temp_dir2)

    try:
        df_excel1 = pd.read_excel(excels1[0], header=None)
        folder_name = str(df_excel1.iloc[4, 2]).strip()
        if not folder_name:
            folder_name = "default_folder"
    except Exception as e:
        print("Error reading C5 cell:", e)
        folder_name = "default_folder"

    folder_name = folder_name.replace("\\", "/").replace(" ", "_")

    base_folder = os.path.join(app.config['UPLOAD_FOLDER'], folder_name)
    folder1 = os.path.join(base_folder, '1')
    folder2 = os.path.join(base_folder, '2')
    os.makedirs(folder1, exist_ok=True)
    os.makedirs(folder2, exist_ok=True)

    for excel in excels1:
        extract_images_from_excel(excel, folder1)
    for excel in excels2:
        extract_images_from_excel(excel, folder2)

    names1 = []
    for excel in excels1:
        names1.extend(read_image_names_from_excel(excel))
    names2 = []
    for excel in excels2:
        names2.extend(read_image_names_from_excel(excel))

    results = []
    for name1, name2 in zip(names1, names2):
        try:
            img1_path = find_file_by_name(folder1, name1)
            img2_path = find_file_by_name(folder2, name2)

            if not img1_path or not img2_path:
                print(f"Skipping missing files: {name1} or {name2}")
                continue

            text1 = perform_ocr(img1_path, language_code)
            text2 = perform_ocr(img2_path, language_code)
            is_match = text1.strip() == text2.strip()

            results.append({
                'image_name': name1,
                'image1_path': f"images/{folder_name}/1/{os.path.basename(img1_path).replace('#', '')}",
                'image2_path': f"images/{folder_name}/2/{os.path.basename(img2_path).replace('#', '')}",
                'text1': text1,
                'text2': text2,
                'is_match': is_match,
                'status': "✅ Match" if is_match else "❌ Mismatch"
            })
        except Exception as e:
            print(f"Error comparing images: {e}")
            continue
        result_data = '||'.join(
    f"{r['image_name']}|{r['text1']}|{r['text2']}|{r['status']}" for r in results
    )
    return render_template('results.html', results=results, result_data=result_data, ask_cleanup=True)

@app.route('/cleanup', methods=['POST'])
def cleanup():
    try:
        shutil.rmtree(TEMP_DIR_1)
        shutil.rmtree(TEMP_DIR_2)
        os.makedirs(TEMP_DIR_1, exist_ok=True)
        os.makedirs(TEMP_DIR_2, exist_ok=True)
        message = "Temp folders cleared successfully."
    except Exception as e:
        message = f"Error clearing temp folders: {e}"
    return render_template('home.html', message=message)
@app.route('/download_results')
def download_results():
    from xlsxwriter import Workbook
    data = request.args.get('data', '')

    output = BytesIO()
    workbook = Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet('OCR Comparison')

    headers = ['Image Name', 'Text 1', 'Text 2', 'Match Status']
    for col, header in enumerate(headers):
        worksheet.write(0, col, header)

    # Highlight format for mismatch
    red_format = workbook.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'})  # Light red fill

    row_num = 1
    for result in data.split('||'):
        if not result.strip():
            continue
        parts = result.split('|')
        if len(parts) == 4:
            for col, val in enumerate(parts):
                if col == 3 and val.strip().lower() == "mismatch":
                    worksheet.write(row_num, col, val, red_format)
                else:
                    worksheet.write(row_num, col, val)
            row_num += 1

    workbook.close()
    output.seek(0)
    return send_file(output,
                     as_attachment=True,
                     download_name="ocr_results.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)

